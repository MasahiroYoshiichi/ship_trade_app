import boto3  # AWS SDKをインポート
import openpyxl  # openpyxlライブラリをインポートしてExcelファイルを処理
import io
from io import BytesIO
import logging  # ロギングの標準ライブラリを追加
import re  # 正規表現を使うためにreモジュールをインポート
import json  # JSONの操作に利用
import traceback  # トレースバックを取得するためにインポート
from datetime import datetime  # 日時の操作に利用
from multipart import parse_form_data
import base64

# ロガーの設定
logger = logging.getLogger()
logger.setLevel(logging.INFO)

# AWSのS3とTextractクライアントを作成
s3_client = boto3.client('s3')
textract_client = boto3.client('textract')

# フォーマット格納S3バケットの指定
format_bucket = 'asian-ad-arrival-notice-format'
format_key = 'linkrun/arrival-notice-format.xlsx'

logger.info(f"format_bucket: {format_bucket}")
logger.info(f"format_key: {format_key}")

# S3からExcelファイルをダウンロードし、openpyxlで開く関数を定義
def download_and_open_excel(format_bucket, format_key):
    # S3 からファイルをダウンロード
    response = s3_client.get_object(
        Bucket=format_bucket,
        Key=format_key
    )
    file_content = response['Body'].read()  # ファイルの内容を読み取る
    # BytesIO オブジェクトにダウンロードした内容を書き込む
    excel_file_io = io.BytesIO(file_content)
    # openpyxl で Excel ファイルを開く
    workbook = openpyxl.load_workbook(excel_file_io)
    return workbook  # Workbookオブジェクトを返す

# 署名付きURLを生成
def generate_presigned_url(bucket, new_key, expiration=3600):
    try:
        response_url = s3_client.generate_presigned_url('get_object',
                                                        Params={'Bucket': bucket,
                                                                'Key': new_key},
                                                        ExpiresIn=expiration)
        return response_url
    except Exception as e:
        logger.error(f"An error occurred: {e}")
        raise


def text_split_text_and_write(sheet, text, start_row, start_col, max_length):
    words = text.split()  # 入力されたテキストを単語に分割する。
    current_line = []  # 現在の行に追加する単語を保持するリストを初期化する。
    row = start_row  # 書き込みを開始する行の番号を設定する。

    for word in words:  # 各単語に対してループを実行する。
        # 現在の行に新しい単語を追加した場合に最大文字数を超えるかどうかを確認する。
        # ' '.join(current_line) + ' ' + word は、現在の行に新しい単語を追加した場合の文字列を作成する。
        if len(' '.join(current_line) + ' ' + word) > max_length:
            # 現在の行をシートに書き込み、新しい行を開始する。
            # ' '.join(current_line) は、現在の行の単語を結合して1つの文字列を作成する。
            sheet.cell(row=row, column=start_col).value = ' '.join(current_line)
            current_line = [word]  # 新しい行を開始し、新しい単語を追加する。
            row += 1  # 次の行に移動する。

        else:
            # 現在の行に単語を追加する。
            # この単語を追加しても最大文字数を超えないことが保証されている。
            current_line.append(word)

    # 残っている単語がある場合は、最後の行を書き込みする。
    # これは、最後の単語が処理された後に残りの単語をシートに書き込むためである。
    if current_line:
        sheet.cell(row=row, column=start_col).value = ' '.join(current_line)
        
def extract_form_data_from_textract_response(response):
    extracted_form_data = {}  # 抽出されたデータを保持するための辞書
    block_id_to_text = {}  # ブロックIDとテキストの対応関係を保持するための辞書

    # まず、すべてのブロックのIDとテキスト情報をblock_id_to_textにマッピングします。
    for block in response['Blocks']:
        if 'Text' in block:
            block_id_to_text[block['Id']] = block['Text']

    for item in response['Blocks']:  # レスポンス内の各ブロックをループ
        if item['BlockType'] == 'KEY_VALUE_SET':  # ブロックタイプがキー値セットの場合
            # テキストを直接取得できる場合はそのテキストを使用し、そうでない場合は空文字を初期値とします。
            text = item.get('Text', '')

            # テキストが直接ブロック内にない場合、関連ブロックを調べてテキスト情報を取得します。
            if not text and 'Relationships' in item:
                for relationship in item['Relationships']:
                    if relationship['Type'] == 'CHILD':
                        for related_id in relationship['Ids']:
                            if related_id in block_id_to_text:
                                text += block_id_to_text[related_id] + ' '

            if 'KEY' in item['EntityTypes']:  # これがキーである場合
                key_name = text.strip()  # キー名を保存
                extracted_form_data[key_name] = ''  # 初期値として空文字を設定
            elif 'VALUE' in item['EntityTypes']:  # これが値である場合
                extracted_form_data[key_name] = text.strip()  # 最後に見たキーに対する値を保存

    return extracted_form_data  # 抽出されたデータを返す

def extract_table_data_from_textract_response(response):
    extracted_table_data = []  # テーブルデータを保持するためのリスト
    block_id_to_text = {}  # ブロックIDとテキストの対応関係を保持するための辞書

    # まず、すべてのブロックのIDとテキスト情報をblock_id_to_textにマッピングします。
    for block in response['Blocks']:
        if 'Text' in block:
            block_id_to_text[block['Id']] = block['Text']


def extract_table_data_from_textract_response(response):
    block_id_to_text = {}  # ブロックIDとテキストの対応関係を保持するための辞書

    # すべてのブロックのIDとテキスト情報をblock_id_to_textにマッピングします。
    for block in response['Blocks']:
        if 'Text' in block:
            block_id_to_text[block['Id']] = block['Text']

    # 最初のテーブルのセルIDを保存するリスト
    first_table_cell_ids = []
    for block in response['Blocks']:
        if block['BlockType'] == 'TABLE':
            for relationship in block.get('Relationships', []):
                if relationship['Type'] == 'CHILD':
                    first_table_cell_ids.extend(relationship['Ids'])
            break  # 最初のテーブルのセルIDを取得したらループを終了

    # セルデータの一時保持
    cell_data = {}
    for item in response['Blocks']:  
        if item['BlockType'] == 'CELL' and item['Id'] in first_table_cell_ids:
            row_index = item['RowIndex'] - 1  
            col_index = item['ColumnIndex'] - 1  
            text = item.get('Text', '')  

            # テキストが直接ブロック内にない場合、関連ブロックを調べてテキスト情報を取得します。
            if not text and 'Relationships' in item:
                child_texts = []
                for relationship in item['Relationships']:
                    if relationship['Type'] == 'CHILD':
                        for related_id in relationship['Ids']:
                            child_texts.append(block_id_to_text.get(related_id, ''))
                text = ' '.join(child_texts).strip()

            cell_data[(row_index, col_index)] = text

    # ヘッダー情報を取得して、それに対応するセルをマッピング
    extracted_table_data = {}
    headers = [cell_data.get((0, j), '') for j in range(len(cell_data.keys()))]
    
    for index, header in enumerate(headers):
        if header:
            extracted_table_data[header] = cell_data.get((1, index), '')

    return extracted_table_data




# Lambdaハンドラ関数を定義
def lambda_handler(event, context):
    try:

        # Event内容をログに出力
        logger.info(f"Event: {json.dumps(event)}")

        # マルチパートフォームデータの解析
        environ = {
            'REQUEST_METHOD': 'POST',
            'CONTENT_TYPE': event['headers']['content-type'],
            'wsgi.input': BytesIO(base64.b64decode(event['body'])),
        }
        form, files = parse_form_data(environ)
        
        
        pdf_data = files['pdf'].file.read()
        
        current_time = datetime.utcnow()
        formatted_time = current_time.strftime('%Y:%m:%d:%H:%M')
        
        path = event['path']
        bl_resource = path.split('/')[-1]
        
        BL_Bucket = 'asian-ad-trade-document'
        # S3キーを生成
        BL_key = f"bill-of-lading/{bl_resource}/{formatted_time}.pdf"
        
        s3_client.put_object(Bucket=BL_Bucket, Key=BL_key, Body=pdf_data)

        # バケット名とオブジェクトキーをログに出力
        logger.info(f"Bucket: {BL_Bucket}")
        logger.info(f"Key: {BL_key}")
        
        # Textractクライアントを使用してドキュメントを分析
        response = textract_client.analyze_document(
            Document={
                'S3Object': {
                    'Bucket': BL_Bucket,
                    'Name': BL_key
                }
            },
            FeatureTypes=["FORMS", "TABLES"]  # フォームとテーブルの特徴を分析
        )
        # Textractのレスポンスをログに出力
        logger.info(f"Textract response: {json.dumps(response)}")

        # Textractのレスポンスからデータを抽出
        form_data = extract_form_data_from_textract_response(response)
        table_data = extract_table_data_from_textract_response(response)
        
        # form_dataとtable_dataの内容をログに出力
        logger.info(f"Form data: {json.dumps(form_data)}")
        logger.info(f"Table data: {json.dumps(table_data)}")

        # Keyとセルをマッピング
        key_to_cell_mapping = {
            "Place of Receipt": "A20",
            "No. of Original B(s)/L": "",
            "Payable at": "",
            "Prepaid at": "",
            "Total Prepaid": "",
            "Port of Discharge": "B24",
            "pre-carriage by": "",
            "Port of Loading": "A24",
            "Place and date of issue": "",
            "Ex. Rate": "",
            "Signed for the Carrier": "",
            "Place of Delivery": "A26",
            "Consignee": "A11",
            "B/L No.": "F22",
            "Ocean Vessel": "B20",
            "Date": "",
            "FREIGHT CHARGES": "D43",
            "Per": "",
            "Voy. No.": "E20",
            "Revenue Tons": "",
            "Rate": "",
            "Collect": "",
            "Shipper": "A5",
            "final Destination for the Merchant's Reference": "",
            "Notify party": "A16",
            "Signature": "",
            "CONTAINERS/SEAL NO.:": "A38",
            "TOTAL NUMBER OF CONTAINERS OR PACKAGES (IN WORDS)": "",
            "Prepaid": "",
            "LADEN ON BOARD THE VESSEL": "",
            "Measurement": "",
            "CY-CY": "",
            "Gross weight": "",
            "PORT OF DISCHARGE": "B24",
            "VESSELIVOYAGE": "E20",
            "PRE CARRIAGE BY": "",
            "PORT OF LOADING": "A24",
            "PLACE OF RECEIPT": "A20",
            "PLACE OF DELIVERY": "A26",
            "PLACE OF ISSUE": "",
            "PREPAID": "",
            "CONSIGNEE/COMPLETE NAME,ADDRESS AND PHONE)": "A11",
            "UNIT": "",
            "EXCESS VALUE DECLARATION": "",
            "COLLECT": "",
            "DATE OF ISSUE": "",
            "RATE": "",
            "NOTIFY PARTY(COMPLETE NAME ADDRESS AND PHONE) I is agreed that no responsibilites sha be attached to the carrier or his agents for failure notify)": "A11",
            "FREIGHT & CHARGES": "",
            "No of Originali Bills of lading": "",
            "TOTAL NUMBER OF CONTAINERS OR PACKAGES(IN WORD)": "",
            "TEMPERATURE CONTROL INSTRUCTION": "",
            "SHIPPERICOMPLETE NAME ADDRESS AND PHONE)": "A5",
            "PREPAIDAT": "",
            "BILL OF LADING NO": "F20",
            "PAYABLE AI": "",
        }
        # ヘッダーとセルのマッピング
        header_to_cell_mapping = {
            'container No.': 'A32',
            'Seal No.: Marks Nos.': 'A35',
            'NO Of containers or pkgs': 'C32',
            'Kind of Packages.Description of Goods': 'D32',
            'Gross weight': 'G33',
            'Measurement': 'H33',
            "CONTAINER NO.SEAL NO MARKS & NUMBERS": "A32",
            "QUANTITY/FOR CUSTOMERS DECLARATION ONLY:": "C32",
            "DESCRIPTION OF GOODS (SAID TO CONTAIN": "D32",
            "GROSS WEIGHT (KILOS)": "G32",
            "MEASURMENT CU METRES)": "H33",
        }
        # S3からExcelファイルをダウンロードし、openpyxlで開く
        workbook = download_and_open_excel(format_bucket, format_key)
        logger.info(f'workbook: {workbook}')
        
        # 処理するシートを取得
        sheet = workbook.active
        for key, cell_address in key_to_cell_mapping.items():  # key_to_cell_mapping辞書をループ処理
            value = form_data.get(key, '')  # キーに対応する値を取得、なければ空文字を使う

            # 列ラベルの正規表現マッチング
            match_result = re.match(r"([A-Za-z]+)", cell_address)
            if not match_result:
                logger.warning(f"Failed to match column label for cell address: {cell_address}. Skipping...")
                continue  # マッチング失敗時はこのイテレーションをスキップ
            col_label = match_result.group(0)  # 列ラベルを抽出
            col_idx = openpyxl.utils.cell.column_index_from_string(col_label)  # 列のインデックスを取得

            # 行インデックスの正規表現マッチング
            search_result = re.search(r"(\d+)", cell_address)
            if not search_result:
                logger.warning(f"Failed to search row index for cell address: {cell_address}. Skipping...")
                continue  # マッチング失敗時はこのイテレーションをスキップ
            row_idx = int(search_result.group(0))  # 行のインデックスを取得

            # split_text_and_write関数を使って、テキストを分割して書き込む
            text_split_text_and_write(sheet, value, row_idx, col_idx, 45)

        for key, cell_address in header_to_cell_mapping.items():  # key_to_cell_mapping辞書をループ処理
            value = table_data.get(key, '')  # キーに対応する値を取得、なければ空文字を使う

            # 列ラベルの正規表現マッチング
            match_result = re.match(r"([A-Za-z]+)", cell_address)
            if not match_result:
                logger.warning(f"Failed to match column label for cell address: {cell_address}. Skipping...")
                continue  # マッチング失敗時はこのイテレーションをスキップ
            col_label = match_result.group(0)  # 列ラベルを抽出
            col_idx = openpyxl.utils.cell.column_index_from_string(col_label)  # 列のインデックスを取得

            # 行インデックスの正規表現マッチング
            search_result = re.search(r"(\d+)", cell_address)
            if not search_result:
                logger.warning(f"Failed to search row index for cell address: {cell_address}. Skipping...")
                continue  # マッチング失敗時はこのイテレーションをスキップ
            row_idx = int(search_result.group(0))  # 行のインデックスを取得

            # split_text_and_write関数を使って、テキストを分割して書き込む
            text_split_text_and_write(sheet, value, row_idx, col_idx, 12)

        # Excelファイルをバイトストリームに保存
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)  # ストリームの位置をリセット

        # 作成したExcelをS3にアップロード。
        # プレフィックスを指定
        new_prefix = 'arrival-notice/'
        # 既存のプレフィックスを新しいプレフィックスに置き換える
        new_key = BL_key.replace('bill-of-lading/', new_prefix).replace('.pdf', '.xlsx')
        logger.info(f"New key: {new_key}")
        # 作成したExcelをS3にアップロード
        s3_client.put_object(Bucket=BL_Bucket, Key=new_key, Body=output.getvalue())
        logger.info(f'Successfully uploaded {new_key} to {BL_Bucket}')
        # 署名付きURLを生成
        presigned_url = generate_presigned_url(BL_Bucket, new_key)
        logger.info(f'Presigned URL: {presigned_url}')
        # 署名付きURLをフロントエンドに返す
        return {
            'statusCode': 200,
            'headers': {
                'Access-Control-Allow-Origin': '*',  
            },
            'body': json.dumps({
                'presignedUrl': presigned_url
            })
        }

    except Exception as e:
        error_message = str(e)
        trace = traceback.format_exc()
        error_timestamp = datetime.utcnow().isoformat()
        error_response = {
            "errorCode": "500",
            "errorMessage": error_message,
            "errorDetail": trace,
            "errorTimestamp": error_timestamp,
            # "recommendedAction": "Contact support or try again later",  # 任意
        }
    # 例外情報をロギング
    logger.exception(f'An error occurred: {error_message}')
    return {
        "statusCode": 500,
        "headers": {
            'Access-Control-Allow-Origin': '*',  
        },
        "body": json.dumps(error_response)
    }

